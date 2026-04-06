from flask import Flask, send_from_directory, jsonify, request
import sqlite3
import os
from openpyxl import load_workbook
from datetime import datetime

app = Flask(__name__)

# DB 경로
BASE_DIR = os.path.dirname(os.path.abspath(__file__))
DB_PATH = os.path.join(BASE_DIR, "attendance.db")

# DB 초기화
def init_db():
    conn = sqlite3.connect(DB_PATH)
    cur = conn.cursor()

    cur.execute("""
    CREATE TABLE IF NOT EXISTS work_schedule (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        name TEXT,
        date TEXT,
        status TEXT
    )
    """)

    cur.execute("""
    CREATE TABLE IF NOT EXISTS special_duty (
        id INTEGER PRIMARY KEY AUTOINCREMENT,
        duty TEXT,
        name TEXT,
        date TEXT
    )
    """)

    conn.commit()
    conn.close()

def get_db():
    conn = sqlite3.connect(DB_PATH)
    conn.row_factory = sqlite3.Row
    return conn

init_db()

@app.route("/schedule")
def get_schedule():
    dates = request.args.get("dates")
    month = request.args.get("month")

    conn = get_db()
    cur = conn.cursor()

    if dates:
        date_list = dates.split(",")
        placeholders = ",".join("?" * len(date_list))
        cur.execute(
            f"SELECT name, date, status FROM work_schedule WHERE date IN ({placeholders})",
            date_list
        )
    elif month:
        cur.execute(
            "SELECT name, date, status FROM work_schedule WHERE date LIKE ?",
            (f"{month}%",)
        )
    else:
        conn.close()
        return jsonify({})

    rows = cur.fetchall()
    conn.close()

    result = {}
    for r in rows:
        result.setdefault(r["name"], {})[r["date"]] = r["status"]

    return jsonify(result)

@app.route("/special")
def get_special():
    month = request.args.get("month")

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "SELECT duty, name, date FROM special_duty WHERE date LIKE ?",
        (f"{month}%",)
    )

    rows = cur.fetchall()
    conn.close()

    result = {}
    for r in rows:
        result.setdefault(r["date"], {}).setdefault(r["duty"], []).append(r["name"])

    return jsonify(result)

@app.route("/add_schedule", methods=["POST"])
def add_schedule():
    data = request.json

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "INSERT INTO work_schedule (name, date, status) VALUES (?, ?, ?)",
        (data["name"], data["date"], data["status"])
    )

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

@app.route("/add_schedule_bulk", methods=["POST"])
def add_schedule_bulk():
    data = request.json["data"]

    conn = get_db()
    cur = conn.cursor()

    for line in data:
        if "\t" in line:
            parts = line.split("\t")
        else:
            parts = line.split(",")

        if len(parts) != 3:
            continue

        name, date, status = [p.strip() for p in parts]

        if not name or not date or not status:
            continue

        cur.execute(
            "INSERT INTO work_schedule (name, date, status) VALUES (?, ?, ?)",
            (name, date, status)
        )

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

@app.route("/add_special", methods=["POST"])
def add_special():
    data = request.json

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "INSERT INTO special_duty (duty, name, date) VALUES (?, ?, ?)",
        (data["duty"], data["name"], data["date"])
    )

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

@app.route("/delete_schedule", methods=["POST"])
def delete_schedule():
    data = request.json

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM work_schedule WHERE name=? AND date=?",
        (data["name"], data["date"])
    )

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

@app.route("/delete_special", methods=["POST"])
def delete_special():
    data = request.json

    conn = get_db()
    cur = conn.cursor()

    cur.execute(
        "DELETE FROM special_duty WHERE name=? AND date=? AND duty=?",
        (data["name"], data["date"], data["duty"])
    )

    conn.commit()
    conn.close()

    return jsonify({"result": "ok"})

@app.route("/clear_schedule", methods=["POST"])
def clear_schedule():
    conn = get_db()
    cur = conn.cursor()
    cur.execute("DELETE FROM work_schedule")
    conn.commit()
    conn.close()
    return jsonify({"result": "ok"})

@app.route("/upload_excel_auto", methods=["POST"])
def upload_excel_auto():
    file = request.files.get("file")

    if not file:
        return jsonify({"message": "파일 없음"}), 400

    wb = load_workbook(file, data_only=True)
    ws = wb.active

    conn = get_db()
    cur = conn.cursor()

    inserted = 0

    # 네 엑셀 양식 기준 고정값
    start_data_row = 5
    date_col = 1  # A열 날짜

    # 열 번호 -> 근무지
    workplace_map = {
        3: "삼릉",
        4: "동남산",
        5: "포석정",
        6: "새갓골",
        7: "용장",
        8: "분소",
    }

    # 이름 사전
    known_names = [
        "남산분소장",
        "김재홍", "강이레", "윤동희",
        "예린", "권용조", "손영인", "옥희영",
        "김영호", "서종명", "고현찬",
        "김복현", "서진숙", "정문길", "김태문",
        "최성복",
        "이성원", "이유형"
    ]

    def split_names(text):
        text = str(text).strip()
        if text in ["", "-", "None", "nan"]:
            return []

        text = text.replace("\n", " ")
        text = text.replace(",", " ")
        text = text.replace("·", " ")
        text = text.replace("/", " ")

        rough_parts = [x.strip() for x in text.split() if x.strip()]
        result = []

        for part in rough_parts:
            if part in known_names:
                result.append(part)
                continue

            temp = part
            matched_any = True

            while temp and matched_any:
                matched_any = False
                for name in sorted(known_names, key=len, reverse=True):
                    if temp.startswith(name):
                        result.append(name)
                        temp = temp[len(name):]
                        matched_any = True
                        break

        return result

    for row in range(start_data_row, ws.max_row + 1):
        raw_date = ws.cell(row=row, column=date_col).value
        if not raw_date:
            continue

        if isinstance(raw_date, datetime):
            date_str = raw_date.strftime("%Y-%m-%d")
        else:
            text_date = str(raw_date).strip()

            if text_date in ["", "None", "nan"]:
                continue

            if "/" in text_date and len(text_date.split("/")) == 2:
                try:
                    month, day = text_date.split("/")
                    date_str = f"2026-{int(month):02d}-{int(day):02d}"
                except:
                    continue
            else:
                date_str = text_date

        for col, workplace in workplace_map.items():
            cell_value = ws.cell(row=row, column=col).value
            if not cell_value:
                continue

            names = split_names(cell_value)

            for name in names:
                cur.execute(
                    "INSERT INTO work_schedule (name, date, status) VALUES (?, ?, ?)",
                    (name, date_str, workplace)
                )
                inserted += 1

    conn.commit()
    conn.close()

    return jsonify({"message": f"{inserted}개 자동 입력 완료"})

@app.route("/")
def index():
    return send_from_directory(os.getcwd(), "test.html")

if __name__ == "__main__":
    app.run(host="0.0.0.0", port=5000)
